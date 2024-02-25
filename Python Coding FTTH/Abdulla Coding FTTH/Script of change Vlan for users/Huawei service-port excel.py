import paramiko
import time
import pandas as pd
import re
from openpyxl import load_workbook

# Define device information
device = {
    "hostname": "10.6.179.243",
    "username": "irqnbn",
    "password": "h5ZotriBozlnop$",
    "port": 22,
}

# Create SSH client object
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

# Establish SSH connection to the device
ssh.connect(**device)

# Create shell object to interact with the device
shell = ssh.invoke_shell()
shell.send('enable \n')
shell.send('config \n')
shell.send('scroll \n')
shell.send(' \n')
time.sleep(1)
shell.send('display current-configuration | include service-port \n')
shell.send(' \n')
time.sleep(20)

# Capture the output from the device
output = b""
while True:
    if shell.recv_ready():
        buffer = shell.recv(1024)
        output += buffer
    else:
        break

# Decode the output
output = output.decode('latin-1')  # Decode using 'latin-1' encoding

# Parse the output and save only the relevant data
lines = output.splitlines()
data = []

# Regular expression pattern to match lines containing "service-port"
pattern = re.compile(r'\b(service-port .+)\b')

for line in lines:
    match = pattern.search(line)
    if match:
        data.append(match.group(1).split())  # Split data into columns

# Create a DataFrame from the data
df = pd.DataFrame(data)

# Create a new DataFrame with only the desired columns B, D, F, and H
selected_columns = df.iloc[:, [1, 3, 5, 7]]  # Columns B, D, F, and H (0-based index)

# Write the selected columns to 
filtered_excel_path = 'filtered_output.xlsx'
# an Excel file
selected_columns.to_excel('filtered_output.xlsx', index=False, header=None)

# Close SSH connection
ssh.close()

# Read the filtered Excel file
filtered_df = pd.read_excel(filtered_excel_path, header=None)

# Extract columns C, D, and E
columns_to_extract = [0, 1, 2 ,3]  # C, D, and E columns (0-based index)
extracted_columns = filtered_df.iloc[:, columns_to_extract]

# Load the existing Excel workbook
excel_path = 'Data Storage.xlsx'
workbook = load_workbook(excel_path)

# Select the existing sheet 'Data'
sheet_name = 'Ddata'
sheet = workbook[sheet_name]

# Determine the next available row in the sheet
next_row = sheet.max_row + 1

# Convert extracted columns to a list
extracted_list = extracted_columns.values.tolist()

# Iterate through the extracted data and append it to the sheet
for row_values in extracted_list:
    sheet.append(row_values)

# Save the modified workbook
workbook.save(excel_path)