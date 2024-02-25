import openpyxl
import pandas as pd
import paramiko
import time
import os
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook


# Create an Excel writer object
writer = pd.ExcelWriter('out.xlsx', engine='xlsxwriter')

# Read Excel file using pandas
df = pd.read_excel(r'D:\Earthlink\New folder\Notes\py\user\FDT1.xlsx', sheet_name="SN")
# Read Excel file using pandas
df2 = pd.read_excel(r'D:\Earthlink\New folder\Notes\py\user\FDT1.xlsx', sheet_name="SN2")

def handle_session(hostname, username, password, port, FDT_Name):
    max_retries = 3  # Maximum number of SSH connection retries
    retry_delay = 5  # Delay between SSH connection retries (in seconds)

    for attempt in range(max_retries):
        try:

            # Create SSH client object
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            # Establish SSH connection
            ssh.connect(hostname=hostname, username=username, password=password, port=port)

            # Create an empty DataFrame to store the dataa
            all_data = pd.DataFrame()

            # create shell object to interact with device
            shell = ssh.invoke_shell()
            shell.send(f"show equipment ont interface \n")

            # wait for command output to be generated
            time.sleep(16)
            
            # Receive the output for display protect-group
            
            output2 = b""
            while True:
                if shell.recv_ready():
                    buffer = shell.recv(1024)
                    output2 += buffer
                else:
                    break
            # Decode using 'latin-1' encoding
            output2 = output2.decode('latin-1')

            # Split the output into lines
            lines2 = output2.splitlines()

            # Parse the output for Sheet2
            data_output2 = [line.split() for line in lines2 if line.strip()]
            time.sleep(1)

            # Create a DataFrame for Sheet2
            df_output2 = pd.DataFrame(data_output2)
            
            # Save the DataFrames for Sheet1 and Sheet2 to Excel
            df_output2.to_excel(writer, index=False)

            #Save the data for the current device in a separate sheet
            all_data.to_excel(writer, index=False)

            #ssh close
            ssh.close()
            break


        except paramiko.AuthenticationException as auth_exception:
            print(f"Authentication failed for {hostname}: {str(auth_exception)}")
            break  # Break the retry loop on authentication failure
        except paramiko.SSHException as ssh_exception:
            print(f"SSH connection failed for {hostname}: {str(ssh_exception)}")
        except Exception as e:
            print(f"An error occurred for {hostname}: {str(e)}")

        # Close SSH connection if it's still open
        if ssh:
            ssh.close()

        if attempt < max_retries - 1:
            print(f"Retrying SSH connection to {hostname} in {retry_delay} seconds...")
            time.sleep(retry_delay)
        else:
            print(f"Max retries exceeded. Could not establish SSH connection to {hostname}.")

# Create a thread pool with a maximum of 1 threads
pool = ThreadPoolExecutor(max_workers=1)

for _, row in df2.iterrows():
    hostname = row['hostname']
    username = 'alifaruq'
    password = 'wnDMOEESqaIsbX9'
    port = '22'
    FDT_Name = row['FDT name']

    pool.submit(handle_session, hostname, username, password, port, FDT_Name)

# Shutdown the thread pool to wait for all tasks to complete
pool.shutdown(wait=True)

# Save the Excel file
writer._save()
time.sleep(2)


# 2 part
# Load the workbooks for both sheets
wb_a = openpyxl.load_workbook('FDT1.xlsx')
wb_b = openpyxl.load_workbook('out.xlsx')

# Assuming data starts from the second row (row index 2)
start_row = 2
# Get the active sheet for both workbooks
FDT1 = wb_a.active
out = wb_b.active
# Step 1: Compare dates in FDT1  (column I) with out (column F)
for row in range(start_row, FDT1.max_row + 1):
    # Get the date value from FDT1, column I
    date_a = FDT1.cell(row=row, column=9).value
    
    # Iterate through each row in out to compare the date
    for b_row in range(start_row, out.max_row + 1):
        # Get the date value from out, column F
        date_b = out.cell(row=b_row, column=6).value
        
        if date_a == date_b:
            # Get the value from out column A
            value_b_col_a = out.cell(row=b_row, column=1).value
            
            # Print out column A value
            print(value_b_col_a)
            
            # Write out column A value into FDT1 column J
            FDT1.cell(row=row, column=10, value=value_b_col_a)
            
            break  # No need to continue searching
           

# Save changes to FDT1
wb_a.save(r'D:\Earthlink\New folder\Notes\py\user\FDT1.xlsx')

# Close the workbooks
wb_a.close()
wb_b.close()
time.sleep(1)

'''
# 3 Part
# Step 2: Print FDT3.xlsx data into FDT1.xlsx sheet "SN"
# Load FDT3.xlsx
wb_a = openpyxl.load_workbook('FDT3.xlsx')
FDT3 = wb_a.active

# Load FDT1.xlsx
wb_c = openpyxl.load_workbook('FDT1.xlsx')
FDT1 = wb_c['SN']  # Assuming "SN" is the sheet name in FDT1.xlsx

# Print FDT3.xlsx data into FDT1.xlsx sheet "SN"
for row in range(start_row, FDT3.max_row + 1):
    # Get the value from FDT3 column J
    value_c_col_j = FDT3.cell(row=row, column=10).value
    
    # Print out column J value into FDT1 sheet "SN" column A
    FDT1.cell(row=row, column=1, value=value_c_col_j)

# Save changes to FDT1.xlsx
wb_a._save('FDT1.xlsx')
time.sleep(1)
'''

# 4 Part
# Create an Excel writer object
writer = pd.ExcelWriter('F.xlsx', engine='xlsxwriter')

# Read Excel file using pandas
df2 = pd.read_excel(r'D:\Earthlink\New folder\Notes\py\user\FDT1.xlsx', sheet_name="SN2")
def handle_session(hostname, username, password, port, FDT_Name,*u_values):
    max_retries = 3  # Maximum number of SSH connection retries
    retry_delay = 5  # Delay between SSH connection retries (in seconds)

    for attempt in range(max_retries):
        try:

            # Create SSH client object
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            # Establish SSH connection
            ssh.connect(hostname=hostname, username=username, password=password, port=port)

            # Create an empty DataFrame to store the dataa
            all_data = pd.DataFrame()

            # create shell object to interact with device
            shell = ssh.invoke_shell()

            # Convert u_values back to a list
            u_values = list(u_values)

            # Use the function to handle U values
            for u_value in u_values:
                # Use the function to handle U values
                shell.send(f"show equipment ont optics {u_value}\n")
                time.sleep(0.5)
            
            time.sleep(2)

            # Receive the output for display protect-group
            output2 = b""
            while True:
                if shell.recv_ready():
                    buffer = shell.recv(1024)
                    output2 += buffer
                else:
                    break
            # Decode using 'latin-1' encoding
            output2 = output2.decode('latin-1')

            # Split the output into lines
            lines2 = output2.splitlines()

            # Parse the output for Sheet2
            data_output2 = [line.split() for line in lines2 if line.strip()]
            time.sleep(2)

            # Create a DataFrame for Sheet2
            df_output2 = pd.DataFrame(data_output2)
            
            # Save the DataFrames for Sheet1 and Sheet2 to Excel
            df_output2.to_excel(writer, index=False)

            #Save the data for the current device in a separate sheet
            all_data.to_excel(writer, index=False)

            #ssh close
            ssh.close()
            break

        except paramiko.AuthenticationException as auth_exception:
            print(f"Authentication failed for {hostname}: {str(auth_exception)}")
            break  # Break the retry loop on authentication failure
        except paramiko.SSHException as ssh_exception:
            print(f"SSH connection failed for {hostname}: {str(ssh_exception)}")
        except Exception as e:
            print(f"An error occurred for {hostname}: {str(e)}")

        # Close SSH connection if it's still open
        if ssh:
            ssh.close()

        if attempt < max_retries - 1:
            print(f"Retrying SSH connection to {hostname} in {retry_delay} seconds...")
            time.sleep(retry_delay)
        else:
            print(f"Max retries exceeded. Could not establish SSH connection to {hostname}.")

# Create a thread pool with a maximum of 1 threads
pool = ThreadPoolExecutor(max_workers=1)
# Usage example
max_u_value = 430  # Change this to the maximum U value you want to query

u_values_list = list(range(1, max_u_value + 1))

for _, row in df2.iterrows():
    hostname = row['hostname']
    username = 'alifaruq'
    password = 'wnDMOEESqaIsbX9'
    port = '22'
    FDT_Name = row['FDT name']
    # Initializing an empty list to store u_values
    u_values = []

    # Using a regular for loop to populate u_values
    for i in range(1, max_u_value + 1):
        u_values.append(row[f'U{i}'])

    pool.submit(handle_session, hostname, username, password, port, FDT_Name,*u_values)

# Shutdown the thread pool to wait for all tasks to complete
pool.shutdown(wait=True)

# Save the Excel file
writer._save()

time.sleep(1)


# 5 Part
# Load the workbooks for both sheets
wb_a = openpyxl.load_workbook('FDT1.xlsx')
wb_b = openpyxl.load_workbook('F.xlsx')

# Assuming data starts from the second row (row index 2)
start_row = 2
# Get the active sheet for both workbooks
FDT1 = wb_a.active
F = wb_b.active
# Step 1: Compare dates in FDT1  (column J) with out (column A)
for row in range(start_row, FDT1.max_row + 1):
    # Get the date value from FDT1, column I
    date_a = FDT1.cell(row=row, column=10).value
    
    # Iterate through each row in out to compare the date
    for b_row in range(start_row, F.max_row + 1):
        # Get the date value from out, column A
        date_b = F.cell(row=b_row, column=1).value
        
        if date_a == date_b:
            # Get the value from out column A
            value_b_col_a = F.cell(row=b_row, column=2).value
            
            # Print out column A value
            print(value_b_col_a)
            
            # Write out column A value into FDT1 column J
            FDT1.cell(row=row, column=11, value=value_b_col_a)
            
            break  # No need to continue searching
           

# Save changes to FDT1
wb_a.save('FDT1.xlsx')

# Close the workbooks
wb_a.close()
wb_b.close()
time.sleep(1)

