#يقرا الداتا من الشيت الاول المسحوب كريشيو لكل الوزرية بكل عامود ويقارنها بالداتا الي بالشيت الثاني المسحوب من الاولتي كيوزرية وبورات ويطلع نتيجة كل يوزر بيا عامود هو وبوراته

import openpyxl

# Load the workbooks for both sheets
wb_a = openpyxl.load_workbook('sheet_a.xlsx')
wb_b = openpyxl.load_workbook('sheet_b.xlsx')

# Assuming data starts from the second row (row index 2)
start_row = 2

# Get the active sheet for both workbooks
sheet_a = wb_a.active
sheet_b = wb_b.active

# Step 1: Compare dates in sheet A (column I) with sheet B (column F)
for row in range(start_row, sheet_a.max_row + 1):
    # Get the date value from sheet A, column I
    date_a = sheet_a.cell(row=row, column=9).value
    
    # Iterate through each row in sheet B to compare the date
    for b_row in range(start_row, sheet_b.max_row + 1):
        # Get the date value from sheet B, column F
        date_b = sheet_b.cell(row=b_row, column=6).value
        
        if date_a == date_b:
            # Get the value from sheet B column A
            value_b_col_a = sheet_b.cell(row=b_row, column=1).value
            
            # Get the value from sheet B column K
            value_b_col_k = sheet_b.cell(row=b_row, column=11).value
            
            # Print sheet B column A value
            print(value_b_col_a)
            
            # Write sheet B column A value into sheet A column Z
            sheet_a.cell(row=row, column=26, value=value_b_col_a)
            
            # Write sheet B column K value into sheet A column AA
            sheet_a.cell(row=row, column=27, value=value_b_col_k)
            break  # No need to continue searching

# Save changes to sheet A
wb_a.save('All_Bg.xlsx')

# Close the workbooks
wb_a.close()
wb_b.close()
