#لاستخراج داتا اليوزرية من الاولتي

import paramiko
import time
import pandas as pd
import re
from openpyxl import load_workbook

# Define device information
device = {
    "hostname": "10.6.177.71",
    "username": "admin",
    "password": "oAoxW0KUsnnl@",
    "port": 22,
}

# Create SSH client object
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

# Establish SSH connection to the device
ssh.connect(**device)

# Create shell object to interact with the device
shell = ssh.invoke_shell()

# Send command to device to set idle-timeout to 1000 sec
shell.send("configure system security ssh server-profile idle-timeout 1000\n")
time.sleep(10)

# Close SSH connection
ssh.close()

# Create SSH client object
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

# Establish SSH connection to the device
ssh.connect(**device)

# Create shell object to interact with the device
shell = ssh.invoke_shell()

# Send command to device to set idle-timeout to 60 sec
shell.send("configure system security ssh server-profile idle-timeout 60\n")
time.sleep(2)
shell.send("exit all\n")
shell.send('show equipment ont interface\n')
time.sleep(20)

# Capture the output from the device
output = b""
while True:
    if shell.recv_ready():
        buffer = shell.recv(1024)
        output += buffer
    else:
        break

# Decode the output
output = output.decode('latin-1')  # Decode using 'latin-1' encoding

# Parse the output and save only the relevant data
lines = output.splitlines()
data = []
start_capture = False  # Flag to indicate when to start capturing data

for line in lines:
    if re.match(r'^\s*interface', line):  # Start capturing when interface table begins
        start_capture = True
        data.append(line.split()[0])  # Add the relevant data to the list
    elif start_capture and not line.strip():  # Stop capturing when an empty line is encountered
        break
    elif start_capture:
        data.append(line.split()[0])  # Add the relevant data to the list

# Create a DataFrame from the data
df = pd.DataFrame({"Column A": data})

# Split the data in 'Column A' based on '/'
split_data = df['Column A'].str.split('/', expand=True)

# Filter the DataFrame based on your conditions (A='1', B='1')
filtered_df = split_data[(split_data[0] == '1') & (split_data[1] == '1')]

# Write the split data to an Excel file
filtered_excel_path = 'filtered_output3.xlsx'
filtered_df.to_excel('filtered_output3.xlsx', index=False, header=None)

# Close SSH connection
ssh.close()

# Read the filtered Excel file
filtered_df = pd.read_excel(filtered_excel_path, header=None)

# Extract columns C, D, and E
columns_to_extract = [2, 3, 4]  # C, D, and E columns (0-based index)
extracted_columns = filtered_df.iloc[:, columns_to_extract]

# Load the existing Excel workbook
excel_path = 'Data Storage.xlsx'
workbook = load_workbook(excel_path)

# Select the existing sheet 'Data'
sheet_name = 'Ddata'
sheet = workbook[sheet_name]

# Determine the next available row in the sheet
next_row = sheet.max_row + 1

# Convert extracted columns to a list
extracted_list = extracted_columns.values.tolist()

# Iterate through the extracted data and append it to the sheet
for row_values in extracted_list:
    sheet.append(row_values)

# Save the modified workbook
workbook.save(excel_path)