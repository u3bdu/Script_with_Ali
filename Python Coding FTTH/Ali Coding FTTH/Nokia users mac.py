#لاستخراج ماكات اليوزرية لاولتيات النوكيا كل اولتي بفايل مستقل


import pandas as pd
import paramiko
import time
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

# Read Excel file using pandas
df = pd.read_excel(r'ipip.xlsx')

def handle_session(hostname, username, password, port, olt_name, seq):
    max_retries = 3  # Maximum number of SSH connection retries
    retry_delay = 5  # Delay between SSH connection retries (in seconds)

    for attempt in range(max_retries):
        try:

            print(olt_name)


            # Create SSH client object
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            # Establish SSH connection
            ssh.connect(hostname=hostname, username=username, password=password, port=port, timeout=120)

            # create shell object to interact with device
            shell = ssh.invoke_shell()

            # send command to device to made idle-timeout 1450 sec
            shell.send("configure system security ssh server-profile idle-timeout 500\n")
            time.sleep(1)
            ssh.close()
            time.sleep(1)
            # Create SSH client object

            # Create SSH client object
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            # Establish SSH connection
            ssh.connect(hostname=hostname, username=username, password=password, port=port)

            # Create an empty DataFrame to store the dataa
            all_data = pd.DataFrame()



            # create shell object to interact with device
            shell = ssh.invoke_shell()
            
            
            # wait for command output to be generated
            
            
            shell.send("show vlan bridge-port-fdb\n")
            time.sleep(400)

            shell.send("configure system security ssh server-profile idle-timeout 60\n")
            time.sleep(1)

            # Receive the output for display current-configuration
            output1 = b""
            while True:
                if shell.recv_ready():
                    buffer = shell.recv(1024)
                    output1 += buffer
                else:
                    break
                

            # Decode using 'latin-1' encoding
            output1 = output1.decode('latin-1')

            # Split the output into lines
            lines1 = output1.splitlines()

            # Parse the output for Sheet1
            data_output1 = [line.split() for line in lines1 if line.strip()]
            time.sleep(2)


            # Create a DataFrame for Sheet1
            df_output1 = pd.DataFrame(data_output1)

            # Get the olt name for the device
            olt_name = df.loc[df['hostname'] == hostname, 'olt name'].values[0]

            # Create a new workbook
            workbook = Workbook()

            # Create worksheets
            sheet1 = workbook.active
            sheet1.title = 'Sheet1'
            sheet2 = workbook.create_sheet(title='Sheet2')

            # Write DataFrame data to worksheets
            for row in dataframe_to_rows(df_output1, index=False, header=False):
                sheet1.append(row)

            # Save the workbook
            workbook.save(f'{olt_name}.xlsx')


           
            #ssh close
            ssh.close()
            break
           


        except paramiko.AuthenticationException as auth_exception:
            print(f"Authentication failed for {hostname}: {str(auth_exception)}")
            break  # Break the retry loop on authentication failure
        except paramiko.SSHException as ssh_exception:
            print(f"SSH connection failed for {hostname}: {str(ssh_exception)}")
        except Exception as e:
            print(f"An error occurred for {hostname}: {str(e)}")

        # Close SSH connection if it's still open
        if ssh:
            ssh.close()

        if attempt < max_retries - 1:
            print(f"Retrying SSH connection to {hostname} in {retry_delay} seconds...")
            time.sleep(retry_delay)
        else:
            print(f"Max retries exceeded. Could not establish SSH connection to {hostname}.")

# Create a thread pool with a maximum of 5 threads
pool = ThreadPoolExecutor(max_workers=20)
 
time.sleep(2) 
for _, row in df.iterrows():
    hostname = row['hostname']
    username = 'karfadhil'
    password = 'bdxSDvjmb940yPD'
    port = '22'
    olt_name = row ['olt name']
    seq = row ['seq']

    try:
        olt_name = df.loc[df['hostname'] == hostname, 'olt name'].values[0]
    except IndexError:
        print(f"Error: 'olt name' not found for {hostname}")
        continue


    pool.submit(handle_session, hostname, username, password, port, olt_name,seq)

# Shutdown the thread pool to wait for all tasks to complete
pool.shutdown(wait=True)